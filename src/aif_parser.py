"""
Parser para hojas AIF (Esquema Ahorro-Inversión-Financiamiento)
del Sector Público Base Caja - Secretaría de Hacienda.

Estructura de columnas (todas las versiones):
  CONCEPTO | TESORO NAC | REC.AFECT | ORG.DESC | SEG.SOC | [EX-CAJAS] | TOTAL | PAMI/FDOS | [TOTAL GRAL]

La posición de las columnas se detecta dinámicamente desde los encabezados.
"""

import re
import io
from datetime import datetime, date

import openpyxl
import xlrd
import pandas as pd


# ──────────────────────────────────────────
# Mapeo de nombres de archivo → (año, mes)
# ──────────────────────────────────────────

MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "sep": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6,
    "jul": 7, "aug": 8, "oct": 10, "nov": 11, "dec": 12,
}


def _year_from_short(yy: str) -> int:
    y = int(yy)
    return 2000 + y if y < 100 else y


def parse_filename_date(filename: str):
    """Extrae (año, mes) del nombre de archivo. Retorna (None, None) si no puede."""
    name = filename.lower().replace("_", " ").replace("-", " ").replace(".", " ")

    # Patrón: mes + 2 o 4 dígitos de año
    for mes_str, mes_num in sorted(MESES_ES.items(), key=lambda x: -len(x[0])):
        pattern = rf"\b{mes_str}\b[^\d]*(\d{{2,4}})\b"
        m = re.search(pattern, name)
        if m:
            return _year_from_short(m.group(1)), mes_num

    # Patrón: año 4 dígitos + mes
    for mes_str, mes_num in sorted(MESES_ES.items(), key=lambda x: -len(x[0])):
        pattern = rf"\b(\d{{4}})\b[^\d]*{mes_str}\b"
        m = re.search(pattern, name)
        if m:
            return int(m.group(1)), mes_num

    # Patrón numérico: 2020 02 → feb 2020
    m = re.search(r"\b(\d{4})\W*(\d{2})\b", name)
    if m:
        return int(m.group(1)), int(m.group(2))

    return None, None


# ──────────────────────────────────────────
# Normalización de conceptos
# ──────────────────────────────────────────

CONCEPTO_NORMALIZE = {
    r"INGRESOS CORRIENTES": "I_INGRESOS_CORRIENTES",
    r"GASTOS CORRIENTES": "II_GASTOS_CORRIENTES",
    r"RESULT.*ECON.*AHORRO": "III_RESULTADO_ECONOMICO",
    r"RECURSOS DE CAPITAL": "IV_RECURSOS_CAPITAL",
    r"GASTOS DE CAPITAL": "V_GASTOS_CAPITAL",
    r"INGRESOS ANTES DE FIGURAT": "VI_INGRESOS_ANTES_FIGURAT",
    r"GASTOS ANTES DE FIGURAT": "VII_GASTOS_ANTES_FIGURAT",
    r"RESULT.*FINANC.*ANTES DE FIGURAT": "VIII_RESULT_FINANCIERO_ANTES_FIGURAT",
    r"CONTRIBUCIONES FIGURATIVAS": "IX_CONTRIBUCIONES_FIGURATIVAS",
    r"GASTOS FIGURATIVOS": "X_GASTOS_FIGURATIVOS",
    r"INGRESOS DESPUES DE FIGURAT": "XI_INGRESOS_DESPUES_FIGURAT",
    r"GASTOS PRIMARIOS DESPUES DE FIGURAT": "XII_GASTOS_PRIMARIOS_DESPUES_FIGURAT",
    r"GASTOS DESPUES DE FIGURAT": "XIII_GASTOS_DESPUES_FIGURAT",
    r"SUPERAVIT PRIMARIO|RESULTADO PRIMARIO": "XIV_RESULTADO_PRIMARIO",
    r"RESULTADO FINANCIERO": "XV_RESULTADO_FINANCIERO",
    # Sub-items relevantes
    r"INGRESOS TRIBUTARIOS|INGRESOS IMPOSITIVOS": "I1_INGRESOS_TRIBUTARIOS",
    # Formato nuevo: "Aportes y contribuciones a la seguridad social"
    # Formato viejo (2021-05/06, 2022-05/07/08): "- Contribuciones a la Seg. Social"
    # IX_CONTRIBUCIONES_FIGURATIVAS ya se matchea antes, no hay colision
    r"APORTES Y CONTRIB.*SEG|CONTRIBUCIONES A LA SEG": "I2_APORTES_SEG_SOCIAL",
    r"INGRESOS NO TRIBUTARIOS|INGRESOS NO IMPOSITIVOS": "I3_INGRESOS_NO_TRIBUTARIOS",
    r"RENTAS DE LA PROPIEDAD": "I4_RENTAS_PROPIEDAD",
    r"TRANSFERENCIAS CORRIENTES.*INGRESOS|INGRESOS.*TRANSFERENCIAS": "I5_TRANSF_CORRIENTES_ING",
    r"GASTOS DE CONSUMO Y OPERACION": "II1_CONSUMO_OPERACION",
    r"REMUNERACIONES": "II1a_REMUNERACIONES",
    r"BIENES Y SERVICIOS": "II1b_BIENES_SERVICIOS",
    r"INTERESES.*RENTAS|INTERESES NETOS": "II2_INTERESES",
    r"PRESTACIONES DE LA SEGURIDAD": "II3_PRESTACIONES_SEG_SOCIAL",
    r"TRANSFERENCIAS CORRIENTES$": "II4_TRANSF_CORRIENTES_GASTO",
    r"AL SECTOR PRIVADO": "II4a_TRANSF_SECTOR_PRIVADO",
    r"AL SECTOR P.BLICO|AL SECTOR PUBLICO": "II4b_TRANSF_SECTOR_PUBLICO",
    # Capital a provincias: ". A Provincias y CABA" (con espacio-A antes de Provincias)
    # Debe ir ANTES que el patron de corrientes para no ser solapado
    r"\. A PROVINCIAS Y CABA| A PROVINCIAS Y CABA": "V2a_TRANSF_CAPITAL_PROVINCIAS",
    # Corrientes a provincias: ".. Provincias y CABA" (sin A)
    r"PROVINCIAS Y CABA": "II4b1_TRANSF_PROVINCIAS_CABA",
    r"UNIVERSIDADES": "II4b2_TRANSF_UNIVERSIDADES",
    r"DEFICIT OPERATIVO EMPRESAS": "II5_DEFICIT_EMPRESAS_PUB",
    r"INVERSION REAL DIRECTA": "V1_INV_REAL_DIRECTA",
    r"TRANSFERENCIAS DE CAPITAL": "V2_TRANSF_CAPITAL",
    r"INVERSION FINANCIERA": "V3_INV_FINANCIERA",
}


def normalize_concepto(raw: str) -> str:
    raw_upper = raw.strip().upper()
    for pattern, code in CONCEPTO_NORMALIZE.items():
        if re.search(pattern, raw_upper):
            return code
    return raw_upper[:80]


# ──────────────────────────────────────────
# Detección de encabezados de columna
# ──────────────────────────────────────────

SUBSECTOR_PATTERNS = {
    "tesoro_nacional": [r"TESORO", r"T\.N\."],
    "rec_afectados": [r"REC\.?\s*AFECT", r"RECURSOS AFECT"],
    "org_descentralizados": [r"ORG\.?\s*DESC", r"ORGANISMOS DESC"],
    "inst_seg_social": [r"SEG\.?\s*SOC", r"SEGURIDAD SOCIAL", r"INST.*SEG"],
    "ex_cajas_prov": [r"EX.CAJAS", r"PVCIALES"],
    "total_adm_nacional": [r"^TOTAL$"],
    "pami_fdos_otros": [r"PAMI", r"FIDUCIARIOS", r"FDOS\."],
    "total_general": [r"T\s*O\s*T\s*A\s*L"],
}


def _match_subsector(cell_text: str) -> str | None:
    t = str(cell_text).strip().upper()
    if not t:
        return None
    for subsector, patterns in SUBSECTOR_PATTERNS.items():
        for p in patterns:
            if re.search(p, t):
                return subsector
    return None


def detect_column_map(rows: list[list]) -> dict:
    """
    Busca las filas de encabezado y devuelve {col_index: subsector_name}.
    Combina hasta 3 filas de encabezado para desambiguar.
    """
    col_map = {}
    # Buscar la fila con "CONCEPTO"
    header_row_idx = None
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if cell and "CONCEPTO" in str(cell).upper():
                header_row_idx = i
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        return col_map

    # Combinar hasta 3 filas de encabezado
    combined = {}
    for delta in range(3):
        ridx = header_row_idx + delta
        if ridx >= len(rows):
            break
        for j, cell in enumerate(rows[ridx]):
            if cell:
                combined[j] = combined.get(j, "") + " " + str(cell).upper()

    # Asignar subsectores
    # Primero pasada: columnas exactas
    assigned = set()
    for j, text in combined.items():
        ss = _match_subsector(text.strip())
        if ss and ss not in assigned:
            col_map[j] = ss
            assigned.add(ss)

    # Segunda pasada para "TOTAL" vs "T O T A L" (si ambos presentes)
    # total_adm_nacional es la primera columna TOTAL, total_general la segunda
    total_cols = [j for j, ss in col_map.items() if ss == "total_general"]
    for j in total_cols:
        # Si hay "EX-CAJAS" columna, este TOTAL es total_adm_nacional
        if "ex_cajas_prov" in col_map.values():
            col_map[j] = "total_adm_nacional"

    return col_map


# ──────────────────────────────────────────
# Detección de fila de datos + título
# ──────────────────────────────────────────

ROMAN_RE = re.compile(r"^(X{0,3}(?:IX|IV|V?I{0,3}))\)", re.IGNORECASE)
STOP_PATTERNS = [r"^\(\d+\)", r"^NOTA", r"^FUENTE", r"^EXCLUYE", r"^INCLUYE"]


def _is_stop_row(cell_text: str) -> bool:
    t = str(cell_text).strip()
    for p in STOP_PATTERNS:
        if re.match(p, t, re.IGNORECASE):
            return True
    return False


def _is_data_row(row: list) -> tuple[bool, str, str]:
    """
    Retorna (is_data, concepto_raw, nivel).
    nivel: 'principal' | 'detalle' | 'subdetalle'
    """
    # La celda de concepto es la primera no vacía de las primeras 3 cols
    concepto_cell = ""
    for col in row[:4]:
        if col is not None and str(col).strip():
            concepto_cell = str(col).strip()
            break

    if not concepto_cell:
        return False, "", ""

    if _is_stop_row(concepto_cell):
        return False, "", ""

    # Detectar nivel por indentación o Roman numeral
    raw = str(row[1] if len(row) > 1 and row[1] else row[0]).strip() if row else ""

    # Columna B (índice 1) como principal portadora del concepto
    # IMPORTANTE: leer el raw SIN strip para conservar espacios de sangría
    col_b_raw = str(row[1]) if len(row) > 1 and row[1] is not None else ""
    col_b = col_b_raw.strip()
    col_a = str(row[0]).strip() if row[0] is not None else ""

    if col_a and ROMAN_RE.match(col_a):
        return True, f"{col_a} {col_b}".strip(), "principal"

    if col_b:
        # Contar espacios en el valor RAW (antes del strip)
        leading_spaces = len(col_b_raw) - len(col_b_raw.lstrip())
        if leading_spaces == 0 and not ROMAN_RE.match(col_b):
            return False, "", ""
        if leading_spaces <= 5:
            nivel = "detalle"
        elif leading_spaces <= 9:
            nivel = "subdetalle"
        else:
            nivel = "micro"
        return True, col_b, nivel

    return False, "", ""


# ──────────────────────────────────────────
# Extracción de título y fecha desde la hoja
# ──────────────────────────────────────────

MES_TITLE_RE = re.compile(
    r"(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)"
    r"[\s\-]+(20\d{2})",
    re.IGNORECASE,
)
ACUM_RE = re.compile(r"ACUMULAD|TRIMES|SEMEST", re.IGNORECASE)


def extract_period_from_sheet(rows: list[list]) -> tuple[int | None, int | None, str]:
    """
    Returns (year, month, tipo) where tipo='mensual'|'acumulado'.
    Scans the first 10 rows for date hints.
    """
    for row in rows[:10]:
        for cell in row:
            if cell is None:
                continue
            s = str(cell)
            m = MES_TITLE_RE.search(s.upper())
            if m:
                mes_num = MESES_ES[m.group(1).lower()]
                year = int(m.group(2))
                tipo = "acumulado" if ACUM_RE.search(s) else "mensual"
                return year, mes_num, tipo
            if ACUM_RE.search(s):
                return None, None, "acumulado"
    return None, None, "mensual"


def extract_date_from_imig_header(rows: list[list]) -> tuple[int | None, int | None]:
    """Para hojas IMIG: busca la celda con la fecha como número Excel o datetime."""
    for row in rows[:10]:
        for cell in row:
            if cell is None:
                continue
            if isinstance(cell, (datetime, date)):
                return cell.year, cell.month
            if isinstance(cell, float) and 40000 < cell < 50000:
                # Excel serial date
                try:
                    from openpyxl.utils.datetime import from_excel
                    d = from_excel(cell)
                    return d.year, d.month
                except Exception:
                    pass
    return None, None


# ──────────────────────────────────────────
# Parser principal de una hoja AIF
# ──────────────────────────────────────────

def parse_aif_sheet(rows: list[list], year: int, month: int, periodo: str, source: str) -> list[dict]:
    """Parsea todas las filas de datos de una hoja AIF."""
    col_map = detect_column_map(rows)
    if not col_map:
        return []

    records = []
    in_data = False

    # Buscar fila de inicio de datos (después del encabezado)
    header_row = None
    for i, row in enumerate(rows):
        for cell in row:
            if cell and "CONCEPTO" in str(cell).upper():
                header_row = i
                break
        if header_row is not None:
            break

    if header_row is None:
        return []

    # Parsear filas de datos
    for row in rows[header_row + 2:]:
        is_data, concepto_raw, nivel = _is_data_row(row)
        if not is_data:
            continue

        concepto_norm = normalize_concepto(concepto_raw)

        for col_idx, subsector in col_map.items():
            if col_idx >= len(row):
                continue
            cell = row[col_idx]
            if cell is None:
                continue
            try:
                val = float(str(cell).replace(",", "."))
            except (ValueError, TypeError):
                continue

            records.append({
                "fecha": f"{year:04d}-{month:02d}" if year and month else None,
                "anio": year,
                "mes": month,
                "periodo": periodo,
                "concepto_codigo": concepto_norm,
                "concepto_descripcion": concepto_raw[:120],
                "concepto_nivel": nivel,
                "subsector": subsector,
                "valor_millones_pesos": val,
                "fuente_archivo": source,
            })

    return records


# ──────────────────────────────────────────
# Lectura de workbook → list[list]
# ──────────────────────────────────────────

def read_xlsx_sheet(data: bytes, sheet_name: str) -> list[list]:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[sheet_name]
    return [list(row) for row in ws.iter_rows(max_row=100, values_only=True)]


def read_xls_sheet(data: bytes, sheet_name: str) -> list[list]:
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheet_by_name(sheet_name)
    result = []
    for i in range(min(100, ws.nrows)):
        row = []
        for j in range(ws.ncols):
            cell = ws.cell(i, j)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    d = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    row.append(d)
                except Exception:
                    row.append(cell.value)
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            else:
                row.append(cell.value)
        result.append(row)
    return result


# ──────────────────────────────────────────
# Identificación de hojas AIF dentro de un workbook
# ──────────────────────────────────────────

AIF_SHEET_KEYWORDS = ["AIF", "AHORRO", "SECTOR_PUB", "ABRIL", "ENERO", "FEBRERO",
                      "MARZO", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE",
                      "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

AIF_CONTENT_MARKERS = ["ESQUEMA AHORRO", "SECTOR PUBLICO BASE CAJA", "INGRESOS CORRIENTES"]


def _sheet_is_aif(rows: list[list]) -> bool:
    for row in rows[:10]:
        for cell in row:
            if cell is None:
                continue
            s = str(cell).upper()
            if any(marker in s for marker in AIF_CONTENT_MARKERS):
                return True
    return False


def _sheet_periodo(sheet_name: str, rows: list[list]) -> str:
    name_up = sheet_name.upper()
    if "ACUM" in name_up:
        return "acumulado"
    for row in rows[:8]:
        for cell in row:
            if cell and ACUM_RE.search(str(cell)):
                return "acumulado"
    return "mensual"


# ──────────────────────────────────────────
# Entry point: parsear un archivo completo
# ──────────────────────────────────────────

def parse_file(filename: str, data: bytes) -> list[dict]:
    """
    Extrae todos los registros AIF de un archivo Excel (xls/xlsx).
    Retorna lista de dicts listos para DataFrame.
    """
    is_xlsx = filename.lower().endswith(".xlsx")
    records = []

    try:
        if is_xlsx:
            wb_meta = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
            sheet_names = wb_meta.sheetnames
            wb_meta.close()
        else:
            wb_meta = xlrd.open_workbook(file_contents=data)
            sheet_names = wb_meta.sheet_names()
    except Exception as e:
        print(f"  [WARN] No se pudo abrir {filename}: {e}")
        return []

    year_fn, month_fn = parse_filename_date(filename)

    for sname in sheet_names:
        try:
            if is_xlsx:
                rows = read_xlsx_sheet(data, sname)
            else:
                rows = read_xls_sheet(data, sname)
        except Exception as e:
            print(f"  [WARN] Error leyendo hoja '{sname}' en {filename}: {e}")
            continue

        if not _sheet_is_aif(rows):
            continue

        year_sh, month_sh, tipo = extract_period_from_sheet(rows)
        year = year_sh or year_fn
        month = month_sh or month_fn
        periodo = _sheet_periodo(sname, rows)

        if not year or not month:
            print(f"  [WARN] No se pudo determinar fecha para {filename} hoja '{sname}'")
            continue

        recs = parse_aif_sheet(rows, year, month, periodo, filename)
        if recs:
            print(f"  [OK] {filename} | {sname} → {year}-{month:02d} ({periodo}) | {len(recs)} registros")
            records.extend(recs)

    return records
