"""
Parser para hojas IMIG (Informe Mensual de Ingresos y Gastos)
del Sector Público Nacional - Secretaría de Hacienda.

Estructura de columnas:
  CONCEPTO (jerarquía en sangría) | ... | valor_actual | valor_anterior

El año/mes de cada columna se detecta desde los encabezados de la hoja.
"""

import io
import re
from datetime import datetime, date

import openpyxl
import xlrd
import pandas as pd


MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "sep": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}

IMIG_CONTENT_MARKERS = ["INGRESOS TOTALES", "GASTOS PRIMARIOS", "RESULTADO PRIMARIO",
                        "INGRESOS Y GASTOS", "IMIG"]

IMIG_STOP_PATTERNS = [r"^\(\*\)", r"^\(\d+\)", r"^NOTAS", r"^FUENTE", r"^BASE CAJA"]


def _is_imig_sheet(rows: list[list]) -> bool:
    for row in rows[:12]:
        for cell in row:
            if cell and any(m in str(cell).upper() for m in IMIG_CONTENT_MARKERS):
                return True
    return False


def _extract_date_from_cell(cell) -> tuple[int, int] | None:
    if isinstance(cell, (datetime, date)):
        return cell.year, cell.month
    if isinstance(cell, float) and 38000 < cell < 52000:
        try:
            import openpyxl.utils.datetime as oxl_dt
            d = oxl_dt.from_excel(cell)
            return d.year, d.month
        except Exception:
            pass
    if isinstance(cell, str):
        m = re.search(r"(20\d{2})[/\-\.](\d{1,2})[/\-\.](\d{1,2})", cell)
        if m:
            return int(m.group(1)), int(m.group(2))
    return None


def detect_value_columns(rows: list[list]) -> list[tuple[int, int, int]]:
    """
    Returns list of (col_idx, year, month) for data columns.
    Scans rows looking for date cells (datetime or excel serial).
    """
    results = []
    for row in rows[:12]:
        for j, cell in enumerate(row):
            if cell is None:
                continue
            parsed = _extract_date_from_cell(cell)
            if parsed:
                yr, mo = parsed
                if 2019 <= yr <= 2030:
                    results.append((j, yr, mo))
    # Remove duplicates preserving order
    seen = set()
    unique = []
    for item in results:
        key = (item[0], item[1], item[2])
        if key not in seen:
            seen.add(key)
            unique.append(item)
    return unique


def _indent_level(cell_text: str) -> int:
    """Detecta nivel de jerarquía por espacios al inicio."""
    s = str(cell_text)
    return len(s) - len(s.lstrip())


def _is_stop(cell_text: str) -> bool:
    t = str(cell_text).strip()
    for p in IMIG_STOP_PATTERNS:
        if re.match(p, t, re.IGNORECASE):
            return True
    return False


CONCEPTO_IMIG_NORMALIZE = {
    r"INGRESOS TOTALES": "INGRESOS_TOTALES",
    r"TRIBUTARIOS": "Tributarios",
    r"IVA": "IVA_neto_reintegros",
    r"GANANCIAS": "Ganancias",
    r"APORTES.*SEGURI": "Aportes_contrib_seg_social",
    r"D.BITOS Y CR.DITOS|DEBITOS Y CREDITOS": "Debitos_creditos",
    r"BIENES PERSONALES": "Bienes_personales",
    r"IMPUESTOS INTERNOS": "Impuestos_internos",
    r"DERECHOS DE EXPORTACI": "Derechos_exportacion",
    r"DERECHOS DE IMPORTACI": "Derechos_importacion",
    r"IMPUESTO PA.S|IMPUESTO PAIS": "Impuesto_PAIS",
    r"RENTAS DE LA PROPIEDAD": "Rentas_propiedad",
    r"FGS COBRADAS": "FGS_cobradas",
    r"OTROS INGRESOS CORRIENTES": "Otros_ingresos_corrientes",
    r"INGRESOS NO TRIBUTARIOS|INGRESOS NO IMPOSITIVOS": "Ingresos_no_tributarios",
    r"INGRESOS DE CAPITAL": "Ingresos_capital",
    r"GASTOS PRIMARIOS": "GASTOS_PRIMARIOS",
    r"GASTOS CORRIENTES PRIMARIOS": "Gastos_corrientes_primarios",
    r"PRESTACIONES SOCIALES": "Prestaciones_sociales",
    r"JUBILACIONES Y PENSIONES": "Jubilaciones_pensiones",
    r"ASIGNACION UNIVERSAL": "AUH",
    r"ASIGNACIONES FAMILIARES": "Asignaciones_familiares",
    r"PENSIONES NO CONTRIBUTIVAS": "Pensiones_no_contributivas",
    r"INSSJP|PAMI": "INSSJP_PAMI",
    r"OTROS PROGRAMAS SOCIALES|OTRAS PROGRAMOS": "Otros_prog_sociales",
    r"SUBSIDIOS ECON": "Subsidios_economicos",
    r"ENERG.A|ENERGIA": "Subsidios_energia",
    r"TRANSPORTE": "Subsidios_transporte",
    r"GASTOS DE FUNCIONAMIENTO": "Gastos_funcionamiento",
    r"SALARIOS": "Salarios",
    r"OTROS GASTOS DE FUNCIONAMIENTO": "Otros_gastos_func",
    r"TRANSFERENCIAS CORRIENTES A PROVINCIAS": "Transf_corrientes_provincias",
    r"TRANSFERENCIAS A UNIVERSIDADES": "Transf_universidades",
    r"OTROS GASTOS CORRIENTES": "Otros_gastos_corrientes",
    r"GASTOS DE CAPITAL": "Gastos_capital",
    r"RESULTADO PRIMARIO": "RESULTADO_PRIMARIO",
    r"INTERESES": "Intereses_netos",
    r"RESULTADO FINANCIERO": "RESULTADO_FINANCIERO",
}


def normalize_concepto_imig(raw: str) -> str:
    raw_upper = raw.strip().upper()
    for pattern, code in CONCEPTO_IMIG_NORMALIZE.items():
        if re.search(pattern, raw_upper):
            return code
    return raw.strip()[:80]


def _find_concepto_col(rows: list[list]) -> int:
    """Busca la columna donde está el concepto (la que contiene 'INGRESOS TOTALES')."""
    for row in rows:
        for j, cell in enumerate(row):
            if cell and "INGRESOS TOTALES" in str(cell).upper():
                return j
    return 1  # default


def parse_imig_sheet(rows: list[list], year: int, month: int,
                     source: str) -> list[dict]:
    """Parsea una hoja IMIG y retorna lista de dicts."""
    val_cols = detect_value_columns(rows)
    concepto_col = _find_concepto_col(rows)

    # Encontrar fila de "INGRESOS TOTALES" como inicio de datos
    start_row = None
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if cell and "INGRESOS TOTALES" in str(cell).upper():
                start_row = i
                break
        if start_row is not None:
            break

    if start_row is None:
        return []

    records = []
    for row in rows[start_row:]:
        concepto_cell = row[concepto_col] if concepto_col < len(row) else None
        if concepto_cell is None:
            continue
        concepto_str = str(concepto_cell).strip()
        if not concepto_str:
            continue
        if _is_stop(concepto_str):
            break

        nivel = _indent_level(str(concepto_cell))
        concepto_norm = normalize_concepto_imig(concepto_str)

        for (col_idx, col_yr, col_mo) in val_cols:
            if col_idx >= len(row):
                continue
            cell_val = row[col_idx]
            if cell_val is None:
                continue
            if isinstance(cell_val, str) and cell_val.strip() in ("#REF!", "#N/A", ""):
                continue
            try:
                val = float(str(cell_val).replace(",", "."))
            except (ValueError, TypeError):
                continue

            records.append({
                "fecha": f"{col_yr:04d}-{col_mo:02d}",
                "anio": col_yr,
                "mes": col_mo,
                "concepto_descripcion": concepto_str[:120],
                "concepto_codigo": concepto_norm,
                "nivel_jerarquia": nivel,
                "valor_millones_pesos": val,
                "fuente_archivo": source,
            })

    return records


def read_xlsx_sheet(data: bytes, sheet_name: str) -> list[list]:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[sheet_name]
    return [list(row) for row in ws.iter_rows(max_row=120, values_only=True)]


def read_xls_sheet(data: bytes, sheet_name: str) -> list[list]:
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheet_by_name(sheet_name)
    result = []
    for i in range(min(120, ws.nrows)):
        row = []
        for j in range(ws.ncols):
            cell = ws.cell(i, j)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode))
                except Exception:
                    row.append(cell.value)
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            else:
                row.append(cell.value)
        result.append(row)
    return result


def parse_file(filename: str, data: bytes) -> list[dict]:
    """Extrae todos los registros IMIG de un archivo Excel."""
    is_xlsx = filename.lower().endswith(".xlsx")
    records = []

    try:
        if is_xlsx:
            wb_meta = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
            sheet_names = wb_meta.sheetnames
            wb_meta.close()
        else:
            wb_meta = xlrd.open_workbook(file_contents=data)
            sheet_names = wb_meta.sheet_names()
    except Exception as e:
        print(f"  [WARN] No se pudo abrir {filename}: {e}")
        return []

    for sname in sheet_names:
        # Skip sheets that are clearly not IMIG data
        sname_up = sname.upper()
        if any(x in sname_up for x in ["ACUMULADO", "MENSUALIZACION", "SALIDA", "VARME"]):
            continue
        try:
            if is_xlsx:
                rows = read_xlsx_sheet(data, sname)
            else:
                rows = read_xls_sheet(data, sname)
        except Exception as e:
            print(f"  [WARN] Error leyendo hoja '{sname}' en {filename}: {e}")
            continue

        if not _is_imig_sheet(rows):
            continue

        val_cols = detect_value_columns(rows)
        if not val_cols:
            continue

        year, month = val_cols[0][1], val_cols[0][2]
        recs = parse_imig_sheet(rows, year, month, filename)
        if recs:
            print(f"  [OK] IMIG {filename} | {sname} → {year}-{month:02d} | {len(recs)} registros")
            records.extend(recs)

    return records
